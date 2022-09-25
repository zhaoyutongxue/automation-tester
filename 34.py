# coding=utf-8
##6.change layout settings back to the same as version1. no full screen
##7.create 2 pages, one of the pages can change the configuration 
##8.add calibration function for with-load situation
##9.add AC and AA button, change default setting to AA
##10.finish AA mode
##11.add popupwindow to calibration button, fix the spreadsheet for AA
##12.fix the calibration for AA
##13. get rid of colorscale for faster response
##14. shcedule colorscale
##15.deactivate the bar code scan function for OE, colorscale it with antoher script.
##16. use that "next job" button to "skip scan" for Elsafe
##17. stop using DPDM offset at noload condition for elsafe, set to 0. Line 63 and made special adjustment for elsafe in line 65
##18. Test button disable after click.
##19. Delete USBAA. Add DAC to control load current. Use Pin7 to chose CC lines.
##20. Create test phases functions(noload&load).
##21. request PD voltages in test phases. verify PD output(9,15,20V) .
##23. show Type C 5V Vcc on column 1 first then swap it with PD 9V.
##24. simplify the logic for showing labels
##25. auto test
##26. fix the label issue, only have 1 label, but change the context so that they won't overlap.
##27. change restart function
##28. try to auto start type C testing. changed DP pass standard to 2.5~2.9V. DM: 1.7~2.3V, fixed colour scale, but did not activate it.
    #replace time delay of CC line to flip(), so that waiting time can be shorter.
##29. add auto test. but didn't activate it
##    to deal with the tuf only works on one side, wait certain time then say one side is not working.
##    add Restart Button, Release Button
##30. add TUF005 function. Test sink codes added.
##32. delete check zero output
##33. set counter to flip()
##34. change test sequence to minimise the change of CC lines. 
##Note that the readings measured via meter are called real readings, find them in calibration

from __future__ import absolute_import, division, print_function, \
                                                   unicode_literals
import tkinter as tk
import time
import ExpanderPi

import csv
import datetime
import time
import openpyxl

import os
import sys
#DAC to control load current
from ExpanderPi import DAC

##import tkinter as tkMessageBox
from tkinter import messagebox
flag = 0

import pdbuddy

win = tk.Tk()
win.attributes('-fullscreen',True)


# create an instance of the ADC        
adc = ExpanderPi.ADC()  
unitsTestedPASS = 0
unitsTestedFAIL = 0

iobus = ExpanderPi.IO()
iobus.set_port_direction(0, 0x00)
iobus.write_port(0, 0x00)
    
# create a DAC instance with the gain set to 1
dac = DAC(1)


##references 
Vcc_Max_noload = 5.8
Vcc_Min_noload = 4.9

Vcc_Max_load = 5.8
Vcc_Min_load = 4.8


DP_MAX = 3.0
DP_MIN = 2.5

DM_MAX = 2.3
DM_MIN = 1.7
#PD reference,noload
PD9V_Vcc_Max_noload = 9.8
PD9V_Vcc_Min_noload = 8
PD15V_Vcc_Max_noload = 15.3
PD15V_Vcc_Min_noload = 13
PD20V_Vcc_Max_noload = 20.3
PD20V_Vcc_Min_noload = 17

#PD reference,load
PD9V_Vcc_Max_load = 9.9
PD9V_Vcc_Min_load = 8.0
PD15V_Vcc_Max_load = 15.3
PD15V_Vcc_Min_load = 13
PD20V_Vcc_Max_load = 20.3
PD20V_Vcc_Min_load = 17

##TypeC_cable_chip = 0.7   #this is the voltage drop across the IDR chip and the plug in cable when full load.
##TypeA_cable_chip = 0.63
noloadDPDMoffset = 0
noloadDPDMoffsetforA = 0.3
Amp_offset_A = 0
Amp_offset_C = 0
TypeC_adc_correct = -0.033 # this is the extra part of voltage that rapsberry pi read. 
TypeA_adc_correct = -0.003

#load the compensation settings 
with open('settings' + '.csv', 'r') as csv_file:
        reader = csv.reader(csv_file)
        for column in reader:
            TypeC_cable_chip=(float(column[0]))
            usbC_DM_cable_chip=(float(column[1]))
            usbC_DP_cable_chip=(float(column[2]))
            
            TypeA_cable_chip=(float(column[3]))
            usbA_DM_cable_chip=(float(column[4]))
            usbA_DP_cable_chip=(float(column[5]))

            
        print('typeC: '+ str(TypeC_cable_chip), str(usbC_DM_cable_chip),
              str(usbC_DP_cable_chip), '\n',
              'typeA: '+str(TypeA_cable_chip), str(usbA_DM_cable_chip),
              str(usbA_DP_cable_chip))






try:
    import ExpanderPi
except ImportError:
    print('Failed to import ExpanderPi from python system path')
    print('Importing from parent folder instead')
    try:
          import sys
          sys.path.append('..')
          import ExpanderPi
    except ImportError:
          raise ImportError(
              "Failed to import library from parent folder")


global usbC_9V1_Vcc_load, usbC_15V1_Vcc_load, usbC_20V1_Vcc_load
global usbC_9V2_Vcc_load, usbC_15V2_Vcc_load, usbC_20V2_Vcc_load





def averageUSBA_noload(n_point=5,t_delta=0.001):
    global usbA_VCCVal_noload, usbA_DMVal_noload, usbA_DPVal_noload
    
    usbA_VCCVal_noload_sam = []; usbA_DMVal_noload_sam = []; usbA_DPVal_noload_sam = []
    
    for i in range(n_point):
        usbA_VCCVal_noload_sam.append(4.0*adc.read_adc_voltage(1, 0)+Amp_offset_A)
        usbA_DPVal_noload_sam.append(2.0*adc.read_adc_voltage(2, 0) + noloadDPDMoffsetforA)
        usbA_DMVal_noload_sam.append(2.0*adc.read_adc_voltage(3, 0)+noloadDPDMoffsetforA)
        time.sleep(t_delta)
        
    usbA_VCCVal_noload = sum(usbA_VCCVal_noload_sam)/ n_point
    usbA_DMVal_noload = sum(usbA_DMVal_noload_sam)/ n_point
    usbA_DPVal_noload = sum(usbA_DPVal_noload_sam)/ n_point

    
    usbA_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbA_VCCVal_noload))
    usbA_DMVal_noloadLabel.config(text='{0:.2f}V'.format(usbA_DMVal_noload))
    usbA_DPVal_noloadLabel.config(text='{0:.2f}V'.format(usbA_DPVal_noload))
        
##    print(usbA_VCCVal_noload_sam)    


def averageUSBA(n_point=5,t_delta=0.001):
    global usbA_VCCVal, usbA_DMVal, usbA_DPVal  
    usbA_VCCVal_sam = []; usbA_DMVal_sam = []; usbA_DPVal_sam = []
    
    for i in range(n_point):
        usbA_VCCVal_sam.append(4.0*adc.read_adc_voltage(1, 0)+TypeA_cable_chip)
        usbA_DPVal_sam.append(2.0*adc.read_adc_voltage(2, 0) + usbA_DP_cable_chip)
        usbA_DMVal_sam.append(2.0*adc.read_adc_voltage(3, 0) + usbA_DM_cable_chip)
        time.sleep(t_delta)
        
    usbA_VCCVal = sum(usbA_VCCVal_sam)/ n_point
    usbA_DMVal = sum(usbA_DMVal_sam)/ n_point
    usbA_DPVal = sum(usbA_DPVal_sam)/ n_point

    
    usbA_VCCLabel.config(text='{0:.2f}V'.format(usbA_VCCVal))
    usbA_DMLabel.config(text='{0:.2f}V'.format(usbA_DMVal))
    usbA_DPLabel.config(text='{0:.2f}V'.format(usbA_DPVal))
  

def averageUSBC1_noload(n_point=5,t_delta=0.001):
    global usbC_VCCVal1_noload
    usbC_VCCVal1_noload_sam= []; usbC_9V1Val_noload_sam = []; usbC_15V1Val_noload_sam =[]; usbC_20V1Val_noload_sam = []
    
    for i in range(n_point):
        usbC_VCCVal1_noload_sam.append(8.0*adc.read_adc_voltage(4, 0)+Amp_offset_C)
        #usbC_9V1Val_noload_sam.append(2.0*adc.read_adc_voltage(5, 0)+noloadDPDMoffset)
        #usbC_15V1Val_noload_sam.append(2.0*adc.read_adc_voltage(6, 0)+noloadDPDMoffset)
        #usbC_20V1Val_noload_sam.append(2.0 * adc.read_adc_voltage(7, 0) + noloadDPDMoffset)
        time.sleep(t_delta)
        
    usbC_VCCVal1_noload = sum(usbC_VCCVal1_noload_sam)/ n_point
    #usbC_9V1Val_noload = sum(usbC_9V1Val_noload_sam)/ n_point
    #usbC_15V1Val_noload = sum(usbC_15V1Val_noload_sam)/ n_point
    #usbC_20V1Val_noload = sum(usbC_20V1Val_noload_sam) / n_point
    print(usbC_VCCVal1_noload)

    usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal1_noload))
    #usbC_9V1Val_noloadLabel.config(text='{0:.2f}V'.format(usbC_9V1Val_noload))
    #usbC_15V1Val_noloadLabel.config(text='{0:.2f}V'.format(usbC_15V1Val_noload))
    #usbC_20V1Val_noloadLabel.config(text='{0:.2f}V'.format(usbC_20V1Val_noload))

def averageUSBC2_noload(n_point=5,t_delta=0.001):
    global usbC_VCCVal2_noload
    usbC_VCCVal2_noload_sam = []; usbC_9V2Val_noload_sam = []; usbC_15V2Val_noload_sam = []; usbC_20V2Val_noload_sam = []

    for i in range(n_point):
        usbC_VCCVal2_noload_sam.append(8.0 * adc.read_adc_voltage(4, 0) + Amp_offset_C)
        #usbC_9V2Val_noload_sam.append(2.0 * adc.read_adc_voltage(5, 0) + noloadDPDMoffset)
        #usbC_15V2Val_noload_sam.append(2.0 * adc.read_adc_voltage(6, 0) + noloadDPDMoffset)
        #usbC_20V2Val_noload_sam.append(2.0 * adc.read_adc_voltage(7, 0) + noloadDPDMoffset)
        time.sleep(t_delta)

    usbC_VCCVal2_noload = sum(usbC_VCCVal2_noload_sam) / n_point
   # usbC_9V2Val_noload = sum(usbC_9V1Val_noload_sam) / n_point
   # usbC_15V2Val_noload = sum(usbC_15V1Val_noload_sam) / n_point
   # usbC_20V2Val_noload = sum(usbC_20V1Val_noload_sam) / n_point

    usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal1_noload))
    #usbC_9V2Val_noloadLabel.config(text='{0:.2f}V'.format(usbC_9V1Val_noload))
    #usbC_15V2Val_noloadLabel.config(text='{0:.2f}V'.format(usbC_15V1Val_noload))
    #usbC_20V2Val_noloadLabel.config(text='{0:.2f}V'.format(usbC_20V1Val_noload))
    print(usbC_VCCVal2_noload)
#USB C 5V output
def averageUSBC1(n_point=5,t_delta=0.001):
    global usbC_VCCVal1_load
    usbC_VCCVal1_sam= []

    
    for i in range(n_point):
        usbC_VCCVal1_sam.append(8.0*(adc.read_adc_voltage(4, 0)))
        time.sleep(t_delta)
        
    usbC_VCCVal1_load = sum(usbC_VCCVal1_sam)/ n_point

    usbC_VCCVal_loadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal1_load))


            
def averageUSBC2(n_point=5,t_delta=0.001):
    global usbC_VCCVal2_load
    usbC_VCCVal2_sam= []
    
    for i in range(n_point):
        usbC_VCCVal2_sam.append(8*(adc.read_adc_voltage(4, 0)))
        time.sleep(t_delta)
        
    usbC_VCCVal2_load = sum(usbC_VCCVal2_sam)/ n_point

    usbC_VCCVal_loadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal2_load))




# take care of PD readings
def averageUSBC_PD(n_point=5, t_delta=0.001):
    global usbC_VCCValPD
    usbC_VCCValPD_sam = [];

    for i in range(n_point):
        usbC_VCCValPD_sam.append(7.0 * adc.read_adc_voltage(4, 0) + TypeC_cable_chip)
        time.sleep(t_delta)

    usbC_VCCValPD = sum(usbC_VCCValPD_sam) / n_point
    print("PD value:")
    print(usbC_VCCValPD)
    #usbC_VCCPDLabel.config(text='{0:.2f}V'.format(usbC_VCCValPD))


class SinkTester:
    """Tests a PD Buddy Sink"""

    cfg_20v = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=20000, vmin=None, vmax=None,i=0,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_20v1A = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=20000, vmin=None, vmax=None,i=1200,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_20v3A = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=20000, vmin=None, vmax=None, i=3000,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_15v = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=15000, vmin=None, vmax=None, i=0,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_15v2A = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=15000, vmin=None, vmax=None, i=1700,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_15v3A = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=15000, vmin=None, vmax=None, i=3000,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_9v = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=9000, vmin=None, vmax=None, i=0,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_9v2A = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=9000, vmin=None, vmax=None, i=2200,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_9v3A = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=9000, vmin=None, vmax=None, i=3000,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_5v = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=5000, vmin=None, vmax=None, i=0,
            idim=pdbuddy.SinkDimension.CURRENT)
    cfg_5v3A = pdbuddy.SinkConfig(status=pdbuddy.SinkStatus.VALID,
            flags=pdbuddy.SinkFlags.NONE, v=5000, vmin=None, vmax=None, i=3000,
            idim=pdbuddy.SinkDimension.CURRENT)

    flip_poll = 0.125
    flip_delay = 1
    

    @staticmethod
    def _get_tty():
        devices = list(pdbuddy.Sink.get_devices())
        
        # If the wrong number of devices is present, complain and exit.
        if len(devices) < 1:
            raise ValueError('No PD Buddy Sink devices found')
        elif len(devices) > 1:
            raise ValueError('Multiple PD Buddy Sink devices found')
        # Set tty to the auto-detected Sink
        return devices[0].device

    def _pre_write(self, cfg):
        """Notify the user that new configuration is being written"""
        print('Writing {v:.1f} V configuration object…'.format(v=cfg.v/1000.))

    def _cfg_verify(self, sink, cfg):
        """Verify that the given configuration is on the Sink"""
        actual_cfg = sink.get_cfg()
        assert actual_cfg == cfg, ('Configuration error; expected config:\n'
                '{cfg}\n'
                'actual config:\n'
                '{actual_cfg}'.format(cfg=cfg, actual_cfg=actual_cfg))

    def _post_write(self, cfg):
        """Notify the user that new configuration has been written"""
        print('Done.')

    def _write_verify(self, sink, cfg):
        """Write a SinkConfig to a Sink and verify that it was written"""
        self._pre_write(cfg)

        sink.set_tmpcfg(cfg)
        sink.write()

        self._cfg_verify(sink, cfg)

        self._post_write(cfg)

    def test(self):
        with pdbuddy.Sink(self._get_tty()) as sink:
            self._test_phase3(sink)


    def _20V(self, sink):
        self._write_verify(sink, self.cfg_20v)

    def _20V1A(self, sink):
        self._write_verify(sink, self.cfg_20v1A)

    def _20V3A(self, sink):
        self._write_verify(sink, self.cfg_20v3A)

    def _15V(self, sink):
        self._write_verify(sink, self.cfg_15v)

    def _15V2A(self, sink):
        self._write_verify(sink, self.cfg_15v2A)

    def _15V3A(self, sink):
        self._write_verify(sink, self.cfg_15v3A)

    def _9V(self, sink):
        self._write_verify(sink, self.cfg_9v)

    def _9V2A(self, sink):
        self._write_verify(sink, self.cfg_9v2A)

    def _9V3A(self, sink):
        self._write_verify(sink, self.cfg_9v2A)

    def _5V(self, sink):
        self._write_verify(sink, self.cfg_5v)
        
    def _5V3A(self, sink):
        self._write_verify(sink, self.cfg_5v3A)
        
    def _disable_output(self,sink):
        sink.output = False

    def _enable_output(self,sink):
        sink.output = True
 

tester = SinkTester()

flip_poll = 0.125
flip_delay = 1
waiting_time = 2
signal_time = 0.1
retest_counter = 0

# this flip() is to detect successful connection.
def flip():
    global usbTestFailed, retest_counter
    counting_time = 0
    
    #Wait for the Sink to be disconnected
    while len(list(pdbuddy.Sink.get_devices())):
        time.sleep(flip_poll)
        print("Sink waiting to be disconnected")
        counting_time += 1
        print(counting_time)
        if counting_time > 30 :
            print("failed to switch CC lines")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text="sink fail to disconnect")
            retest_counter += 1
            #try to connect for 3 times. 
            if retest_counter < 3 :
                Retest()
            else:
                counting_time = 0
                usbTestFailed = 1
                testButton.config(bg="red", text="TEST")
                RetestButton.config(bg="red", text="RETEST")
                label_passLabel.config(bg="red", text="FAIL")
                Release_buttons()
            break
         
    print("Sink disconnected")

    
    #Wait for the Sink to be connected
    while not len(list(pdbuddy.Sink.get_devices())):
        time.sleep(flip_poll)
        print("Sink waiting to be connected")
        counting_time += 1
        print(counting_time)
        if counting_time > 30 :
            print("one side of Type C not working")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text="type C fail on one side")
            testButton.config(bg="red", text="TEST")
            RetestButton.config(bg=orig_color, text="RETEST")
            label_passLabel.config(bg="red", text="FAIL")
            Release_buttons()
            break
    

    # popupvalue = messagebox.askquestion('WARNING!!!', 'Do you want to restart the program?')
    # if popupvalue == 'yes':
    #     os.system("python3 /home/pi/29.py&")
    #     win.destroy()


    if not usbTestFailed == 1:
        print("Sink connected")
        time.sleep(flip_delay)

#no difference between TUF004 and TUF005 when there is no load.
def test_phase_CC1():
    
    global usbC_VCCValPD, usbC_VCCVal1_noload
    global usbC_9V1_Vcc_noload, usbC_15V1_Vcc_noload, usbC_20V1_Vcc_noload
    global usbC_9V1_Vcc_load, usbC_15V1_Vcc_load, usbC_20V1_Vcc_load
    global usbC_VCCVal0_noload, usbC_VCCVal00_noload
    global waiting_time, signal_time
    
##start NOOOOOOO load voltage checking, set both DAC channels to 0, so that current is 0.
    #DAC1 for Type A and DAC2 for Type C.
    dac.set_dac_voltage(1, 0)
    dac.set_dac_voltage(2, 0)
    time.sleep(0.1)

    ##Check TypeA no load voltages
    print("start to test TypeA noload")
    averageUSBA_noload()
    ##Check TypeC no load voltages

    # 3. Use IO pin 7 to send logic signal to VL.  High signal represents CC1.
    print("start to test TypeC noload, CC1")
    iobus.write_pin(7, 1)
    flip()
    with pdbuddy.Sink(tester._get_tty()) as sink:
        #1.Make sure output is disabled to start
        sink.output = False
        time.sleep(signal_time)
        #2.Check output voltage=0 first time
        # store the result for 0V
        # averageUSBC_PD()
        # usbC_VCCVal0_noload = usbC_VCCValPD
        # usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal0_noload))
        #3Check output at 5V (CC1)
        tester._5V(sink)
        time.sleep(signal_time)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 5V
        averageUSBC1_noload()

        #4.Check output voltage=0 second time
        sink.output = False
        time.sleep(waiting_time)
        # store the result for 0V
        # averageUSBC_PD()
        # usbC_VCCVal00_noload = usbC_VCCValPD
        # usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal00_noload))



        #4. check 9V 15V 20V
        # check 9V
        tester._9V(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 9V
        averageUSBC_PD()
        usbC_9V1_Vcc_noload = usbC_VCCValPD
        usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_9V1_Vcc_noload))


        # check 15V
        tester._15V(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store result for 15V
        averageUSBC_PD()
        usbC_15V1_Vcc_noload = usbC_VCCValPD
        usbC_15V_noloadLabel.config(text='{0:.2f}V'.format(usbC_15V1_Vcc_noload))

        # check 20V
        tester._20V(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 20V
        averageUSBC_PD()
        usbC_20V1_Vcc_noload = usbC_VCCValPD
        usbC_20V_noloadLabel.config(text='{0:.2f}V'.format(usbC_20V1_Vcc_noload))
        # sink.output = False
        # sink.erase()
    print("start to test type C-load, CC1 ")

    if TUFflag == 1:
        PDBloadTUF004CC1()
        # set type A to 2amps:
        dac.set_dac_voltage(1, 0.2)
        # set type C to 2amps:
        dac.set_dac_voltage(2, 0.2)
    else:
        PDBloadTUF005CC1()
        # set type A to 2amps:
        dac.set_dac_voltage(1, 0.2)
        # set type C to 3amps:
        dac.set_dac_voltage(2, 0.3)
    # set type A type C to 0 amps(shut down):
    dac.set_dac_voltage(1, 0)
    dac.set_dac_voltage(2, 0)



# load current is different for TUF004 and TUF005
def test_phase_CC2():
    global usbC_VCCVal, usbC_VCCVal0, usbC_VCCValPD
    global usbC_9V2_Vcc_noload, usbC_15V2_Vcc_noload, usbC_20V2_Vcc_noload
    global usbC_9V2_Vcc_load, usbC_15V2_Vcc_load, usbC_20V2_Vcc_load
    global signal_time, waiting_time
    
##start with load, turn on both A and C resistor
    # test A first, but make sure C is on in order to test maximum 5A output power.
    #TypeA: 2Amps         Type C: 3Amps
    #Type A: DAC1         Type C: DAC2

    waiting_time = 2
    signal_time = 0.1
    ##Check TypeA load voltages
    #Type A test : Turn on CC1, so we can test A when C is on. Type C will take 3A 5V, so total 15W. Type A take 2A, so charger on its max output.
    #Current controlled by op amp need to be calibrated correctly.
    #print("start to test type C with load ")


    ##check Type C, shut down Type A. Check PD: 5V3A, 9V2.22A, 15V1.7A, 20V1.2A.

    # set type A type C to 0 amps(shut down):
    dac.set_dac_voltage(1, 0)
    dac.set_dac_voltage(2, 0)


    ############################CC2
    # 5.Use IO pin 7 to send logic signal to VL.  Low signal represents CC2.
    print("start to test TypeC noload, CC2")
    iobus.write_pin(7, 0)
    flip()
    with pdbuddy.Sink(tester._get_tty()) as sink:
        # 5.Check output voltages when CC2 connected
        # 3Check output at 5V (CC2)
        tester._5V(sink)
        sink.output = True
        time.sleep(waiting_time)
        averageUSBC2_noload()

        # 6.Check output voltage=0 third time
        sink.output = False
        time.sleep(waiting_time)
        usbC_VCCVal00_noload = 7.0 * adc.read_adc_voltage(4, 0)
        usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal00_noload))

        # 7. check 9V 15V 20V
        # check 9V
        tester._9V(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 9V
        averageUSBC_PD()
        usbC_9V2_Vcc_noload = usbC_VCCValPD
        usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_9V2_Vcc_noload))

        # check 15V
        tester._15V(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store result for 15V
        averageUSBC_PD()
        usbC_15V2_Vcc_noload = usbC_VCCValPD
        usbC_15V_noloadLabel.config(text='{0:.2f}V'.format(usbC_15V2_Vcc_noload))

        # check 20V
        tester._20V(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 20V
        averageUSBC_PD()
        usbC_20V2_Vcc_noload = usbC_VCCValPD
        usbC_20V_noloadLabel.config(text='{0:.2f}V'.format(usbC_20V2_Vcc_noload))
        sink.output = False


    print("start to test type C-load CC2")
    if TUFflag == 1:
        PDBloadTUF004CC2()
        # set type A to 2amps:
        dac.set_dac_voltage(1, 0.2)
        # set type C to 2amps:
        dac.set_dac_voltage(2, 0.2)
    else:
        PDBloadTUF005CC2()
        # set type A to 2amps:
        dac.set_dac_voltage(1, 0.2)
        # set type C to 3amps:
        dac.set_dac_voltage(2, 0.3)

    time.sleep(signal_time)
    # #you may delete this code as it's already on 5V
    # with pdbuddy.Sink(tester._get_tty()) as sink:
    #     sink.output = True
    #     tester._5V3A(sink)
    print("start to test type A-load ")
    #test typeA
    averageUSBA()
    # set type A type C to 0 amps(shut down):
    dac.set_dac_voltage(1, 0)
    dac.set_dac_voltage(2, 0)

def PDBloadTUF004CC1():
    global usbC_VCCVal, usbC_VCCVal0, usbC_VCCValPD
    global usbC_9V1_Vcc_load, usbC_15V1_Vcc_load, usbC_20V1_Vcc_load
    global waiting_time, signal_time
    with pdbuddy.Sink(tester._get_tty()) as sink:

        # 2.Check output voltage=0, first time
        sink.output = False
        time.sleep(signal_time)
        # usbC_VCCVal = 7.0 * adc.read_adc_voltage(4, 0)


        # 3Check output at 5V (CC1), 2A，Apple 2A? OR 5V3A
        dac.set_dac_voltage(2, 0.3)
        tester._5V3A(sink)
        time.sleep(signal_time)
        sink.output = True
        time.sleep(waiting_time)
        averageUSBC1()
        # 4.Check output voltage=0, second time
        sink.output = False
        time.sleep(signal_time)
        # usbC_VCCVal0 = 7.0 * adc.read_adc_voltage(4, 0)


        # 4. check 9V 15V 20V

        # check 9V2A
        dac.set_dac_voltage(2, 0.2)
        tester._9V2A(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 9V
        averageUSBC_PD()
        usbC_9V1_Vcc_load = usbC_VCCValPD
        usbC_VCCVal_loadLabel.config(text='{0:.2f}V'.format(usbC_9V1_Vcc_load))

        # check 15V1.7A
        dac.set_dac_voltage(2, 0.17)
        tester._15V2A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store result for 15V1.7A
        averageUSBC_PD()
        usbC_15V1_Vcc_load = usbC_VCCValPD
        usbC_15V_loadLabel.config(text='{0:.2f}V'.format(usbC_15V1_Vcc_load))

        # check 20V1.2A
        dac.set_dac_voltage(2, 0.12)
        tester._20V1A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store the result for 20V
        averageUSBC_PD()
        usbC_20V1_Vcc_load = usbC_VCCValPD
        usbC_20V_loadLabel.config(text='{0:.2f}V'.format(usbC_20V1_Vcc_load))
        time.sleep(waiting_time)

def PDBloadTUF004CC2():
    global usbC_VCCVal, usbC_VCCVal0, usbC_VCCValPD
    global usbC_9V2_Vcc_load, usbC_15V2_Vcc_load, usbC_20V2_Vcc_load
    global waiting_time, signal_time
    with pdbuddy.Sink(tester._get_tty()) as sink:

        # 2.Check output voltage=0, third time
        sink.output = False
        time.sleep(signal_time)
        # usbC_VCCVal00_noload = 7.0 * adc.read_adc_voltage(4, 0)
        # usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal00_noload))

        # 3.Check 5V output voltages when CC2 connected,5V3A
        dac.set_dac_voltage(2, 0.3)
        tester._5V3A(sink)
        time.sleep(signal_time)
        sink.output = True
        time.sleep(waiting_time)
        averageUSBC2()

        # 4.1Check output voltage=0
        sink.output = False
        time.sleep(signal_time)
        # usbC_VCCVal00_noload = 7.0 * adc.read_adc_voltage(4, 0)
        # usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal00_noload))

        # 5. check 9V 15V 20V
        # check 9V2.22A
        dac.set_dac_voltage(2, 0.2)
        tester._9V2A(sink)
        time.sleep(signal_time)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 9V
        averageUSBC_PD()
        usbC_9V2_Vcc_load = usbC_VCCValPD
        usbC_VCCVal_loadLabel.config(text='{0:.2f}V'.format(usbC_9V2_Vcc_load))

        # check 15V1.77A
        dac.set_dac_voltage(2, 0.17)
        tester._15V2A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store result for 15V
        averageUSBC_PD()
        usbC_15V2_Vcc_load = usbC_VCCValPD
        usbC_15V_loadLabel.config(text='{0:.2f}V'.format(usbC_15V2_Vcc_load))
        # check 20V1.2A
        dac.set_dac_voltage(2, 0.12)
        tester._20V1A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store the result for 20V
        averageUSBC_PD()
        usbC_20V2_Vcc_load = usbC_VCCValPD
        usbC_20V_loadLabel.config(text='{0:.2f}V'.format(usbC_20V2_Vcc_load))
        sink.output = False
        time.sleep(signal_time)

        #6.Shutdown everything
        sink.output = False
        dac.set_dac_voltage(1, 0)
        dac.set_dac_voltage(2, 0)
        sink.erase()



# change the following function to 3A TUF005
def PDBloadTUF005CC1():
    global usbC_VCCVal, usbC_VCCVal0, usbC_VCCValPD
    global usbC_9V1_Vcc_load, usbC_15V1_Vcc_load, usbC_20V1_Vcc_load
    global waiting_time, signal_time

    with pdbuddy.Sink(tester._get_tty()) as sink:

        # 2.Check output voltage=0, first time
        sink.output = False
        time.sleep(signal_time)
        usbC_VCCVal = 7.0 * adc.read_adc_voltage(4, 0)


        # 3Check output at 5V (CC1), 2A，Apple 2A? OR 5V3A
        dac.set_dac_voltage(2, 0.3)
        tester._5V3A(sink)
        sink.output = True
        time.sleep(waiting_time)
        averageUSBC1()
        # 4.Check output voltage=0, second time
        sink.output = False
        time.sleep(signal_time)
        usbC_VCCVal0 = 7.0 * adc.read_adc_voltage(4, 0)


        # 4. check 9V 15V 20V

        # check 9V3A
        dac.set_dac_voltage(2, 0.3)
        tester._9V3A(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 9V
        averageUSBC_PD()
        usbC_9V1_Vcc_load = usbC_VCCValPD
        usbC_VCCVal_loadLabel.config(text='{0:.2f}V'.format(usbC_9V1_Vcc_load))

        # check 15V3A
        dac.set_dac_voltage(2, 3)
        tester._15V3A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store result for 15V1.7A
        averageUSBC_PD()
        usbC_15V1_Vcc_load = usbC_VCCValPD
        usbC_15V_loadLabel.config(text='{0:.2f}V'.format(usbC_15V1_Vcc_load))

        # check 20V3A
        dac.set_dac_voltage(2, 3)
        tester._20V3A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store the result for 20V
        averageUSBC_PD()
        usbC_20V1_Vcc_load = usbC_VCCValPD
        usbC_20V_loadLabel.config(text='{0:.2f}V'.format(usbC_20V1_Vcc_load))
        time.sleep(waiting_time)

def PDBloadTUF005CC2():
    global usbC_VCCVal, usbC_VCCVal0, usbC_VCCValPD
    global usbC_9V2_Vcc_load, usbC_15V2_Vcc_load, usbC_20V2_Vcc_load
    global waiting_time, signal_time

    with pdbuddy.Sink(tester._get_tty()) as sink:

        # 2.Check output voltage=0, third time
        sink.output = False
        time.sleep(signal_time)
        usbC_VCCVal00_noload = 7.0 * adc.read_adc_voltage(4, 0)
        usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal00_noload))

        # 3.Check 5V output voltages when CC2 connected,5V3A
        dac.set_dac_voltage(2, 0.3)
        tester._5V3A(sink)
        sink.output = True
        time.sleep(waiting_time)
        averageUSBC2()

        # 4.1Check output voltage=0
        sink.output = False
        time.sleep(signal_time)
        usbC_VCCVal00_noload = 7.0 * adc.read_adc_voltage(4, 0)
        usbC_VCCVal_noloadLabel.config(text='{0:.2f}V'.format(usbC_VCCVal00_noload))

        # 5. check 9V 15V 20V
        # check 9V3A
        dac.set_dac_voltage(2, 0.2)
        tester._9V3A(sink)
        sink.output = True
        time.sleep(waiting_time)
        # store the result for 9V
        averageUSBC_PD()
        usbC_9V2_Vcc_load = usbC_VCCValPD
        usbC_VCCVal_loadLabel.config(text='{0:.2f}V'.format(usbC_9V2_Vcc_load))

        # check 15V3A
        dac.set_dac_voltage(2, 0.17)
        tester._15V3A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store result for 15V
        averageUSBC_PD()
        usbC_15V2_Vcc_load = usbC_VCCValPD
        usbC_15V_loadLabel.config(text='{0:.2f}V'.format(usbC_15V2_Vcc_load))
        # check 20V3A
        dac.set_dac_voltage(2, 0.12)
        tester._20V3A(sink)
        #sink.output = True
        time.sleep(waiting_time)
        # store the result for 20V
        averageUSBC_PD()
        usbC_20V2_Vcc_load = usbC_VCCValPD
        usbC_20V_loadLabel.config(text='{0:.2f}V'.format(usbC_20V2_Vcc_load))
        sink.output = False

        #6.Shutdown everything
        sink.output = False
        dac.set_dac_voltage(1, 0)
        dac.set_dac_voltage(2, 0)
        sink.erase()



def AUTOTEST():
    # if Skip scan button has been pressed, or Barcode has been scanned, this function will start
    # 2 criteria: PD buddy has been detected and Type A has reading above 5V.
    global usbA_VCCVal

    dac.set_dac_voltage(1, 0)
    dac.set_dac_voltage(2, 0)
    time.sleep(0.1)

    usbA_VCCVal = 4.0 * adc.read_adc_voltage(1, 0)

    if usbA_VCCVal >= 4.9:
        print("Type A voltage detected")
        qcTestUSB()
    else:
        rolling_displayLabel.config(bg="blue", text="waiting to start")

    print("start to monitor type A voltage")


# For normal test, it will go through main test, and write result to excel and then release button.
# But for retest, result will be wrote differently
def qcTestUSB():

    MainTest()
    WriteExcelforTest()


    if usbTestFailed > 0 :
        popupreturn_test = messagebox.askretrycancel('Tested fail', 'Readings WRONG!! Press Retry button to Retest.')
        if popupreturn_test is True:
            print("return value:")
            print(popupreturn_test)
            Retest()
            print("back to test lines")

    CountTestNumber()
    Release_buttons()

# if normal test fail, user can use retest function.
# pop up different window
def Retest():
    MainTest()
    WriteExcelforRetest()
    Release_buttons()

def CountTestNumber():
    global unitsTestedFAIL, unitsTestedPASS, test_number

    if (usbTestFailed > 0):
        unitsTestedFAIL += 1
    else:
        unitsTestedPASS += 1

    test_number = unitsTestedFAIL + unitsTestedPASS

    infoLabelString = "TESTED " + str(unitsTestedFAIL + unitsTestedPASS) + " PASS " + str(
        unitsTestedPASS) + " FAIL " + str(unitsTestedFAIL)
    infoLabel.config(text=infoLabelString)



###starting normal test.  
def MainTest():
    global unitsTestedPASS, unitsTestedFAIL, test_number, test_result
    global usbA_VCCVal, usbA_DMVal, usbA_DPVal
    global usbA_VCCVal_noload, usbA_DMVal_noload, usbA_DPVal_noload

    #type C variables:
    global usbC_VCCzero_noload, usbC_VCCzero_load
    global usbC_VCCVal1_noload, usbC_9V1_Vcc_noload, usbC_15V1_Vcc_noload, usbC_20V1_Vcc_noload
    global usbC_VCCVal2_noload, usbC_9V2_Vcc_noload, usbC_15V2_Vcc_noload, usbC_20V2_Vcc_noload
    global usbC_VCCVal1_load, usbC_9V1_Vcc_load, usbC_15V1_Vcc_load, usbC_20V1_Vcc_load
    global usbC_VCCVal2_load, usbC_9V2_Vcc_load, usbC_15V2_Vcc_load, usbC_20V2_Vcc_load

    global A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,result
    global flag, wb, ws, reading_time, usbTestFailed,unitsTestedFAIL,unitsTestedPASS, test_number
    #18. void button in global so that we can call it in functions and then change their configs
    global testButton, RetestButton
#to displas PD pass or not on right bottom corner
    global displayPDfail_9V 

# detect whether someone scan the barcode
    if (flag == 0 ):
        print("did not scan")
        messagebox.askokcancel('Reminder','It seems like you forgot to scan the barcode.')
        return;
    print('flag: ' + str(flag))
    
   
# 18. (append the button via config.so it will be disabled after click) 
    testButton.config(state=tk.DISABLED)
    RetestButton.config(state=tk.DISABLED)
    print("TEST button pressed")
    
    
        
    usbA_VCCVal = 0
    usbA_DMVal = 0
    usbA_DPVal = 0
    usbA_VCCVal_noload = 0
    usbA_DMVal_noload = 0
    usbA_DPVal_noload = 0
    
    # check zero output
    usbC_VCCzero_noload = 1
    usbC_VCCzero_load = 1

    usbC_VCCVal0 = 0
    usbC_VCCVal00 = 0
    usbC_VCCVal000 = 0
 
    #check 5V,noload
    usbC_VCCVal1_noload = 0
    usbC_VCCVal2_noload = 0
    #check PD, noload
    usbC_9V1_Vcc_noload = 0
    usbC_9V2_Vcc_noload = 0
    usbC_15V1_Vcc_noload = 0
    usbC_15V2_Vcc_noload = 0
    usbC_20V1_Vcc_noload = 0
    usbC_20V2_Vcc_noload = 0

    #check 5V,load
    usbC_VCCVal1 = 0
    usbC_VCCVal2 = 0
    usbC_VCCVal1_load = 0
    usbC_VCCVal2_load = 0
    #check PD, load
    usbC_9V1_Vcc_load = 0
    usbC_9V2_Vcc_load = 0
    usbC_15V1_Vcc_load = 0
    usbC_15V2_Vcc_load = 0
    usbC_20V1_Vcc_load = 0
    usbC_20V2_Vcc_load = 0

#display blue color while testing 
    testButton.config(bg="blue", text ="Testing")
    RetestButton.config(bg="blue", text ="Testing")
    label_passLabel.config(bg="blue", text ="Testing")
    #type A label
    usbA_VCCVal_noloadLabel.config(bg="blue")
    usbA_DMVal_noloadLabel.config(bg="blue")
    usbA_DPVal_noloadLabel.config(bg="blue")
    usbA_VCCLabel.config(bg="blue")
    usbA_DMLabel.config(bg="blue")
    usbA_DPLabel.config(bg="blue")

    #type C label
    usbC_VCCVal_noloadLabel.config(bg="blue")
    usbC_15V_noloadLabel.config(bg="blue")
    usbC_20V_noloadLabel.config(bg="blue")
    usbC_VCCVal_loadLabel.config(bg="blue")
    usbC_15V_loadLabel.config(bg="blue")
    usbC_20V_loadLabel.config(bg="blue")

#display blue and all zero
    win.update();

    
    usbTestFailed = 0
        
### no load test
    

    for i in range(1):
        test_phase_CC1()

        if (usbA_VCCVal_noload >=  Vcc_Min_noload) and  (usbA_VCCVal_noload <= Vcc_Max_noload):
            usbA_VCCVal_noloadLabel.config(bg = "green")           
        else:
            usbA_VCCVal_noloadLabel.config(bg = "red")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text ="usbA_VCC fail,no load");
            
        if (usbA_DMVal_noload >= DM_MIN) and  (usbA_DMVal_noload <= DM_MAX):
            usbA_DMVal_noloadLabel.config(bg = "green")
        else:
            usbA_DMVal_noloadLabel.config(bg = "red")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text ="usbA_DM fail,no load")

        if (usbA_DPVal_noload >= DP_MIN) and  (usbA_DPVal_noload <= DP_MAX):
            usbA_DPVal_noloadLabel.config(bg = "green")
        else:
            usbA_DPVal_noloadLabel.config(bg = "red")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text ="usbA_DP fail,no load")

        #start to write Type C codes, no load


        #check 5V,noload
        if (usbC_VCCVal1_noload > Vcc_Min_noload) and  (usbC_VCCVal1_noload <= Vcc_Max_noload):
            if (usbC_VCCVal2_noload > Vcc_Min_noload) and (usbC_VCCVal2_noload <= Vcc_Max_noload):
                usbC_VCCVal_noloadLabel.config(bg="green")
                print("pass")
            else:
                usbC_VCCVal_noloadLabel.config(bg="red", text='{0:.2f}V'.format(usbC_VCCVal2_noload))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 fail_noload")
                print("Type CCCCCCCCCCCC Vcc2 fail")
                usbTestFailed = 1;

        else:
            usbC_VCCVal_noloadLabel.config(bg="red", text='{0:.2f}V'.format(usbC_VCCVal1_noload))
            rolling_displayLabel.config(bg="red", text="usbC_VCC1 fail_noload")

            usbTestFailed = 1;
        print("USB C values!!!!!! ")
        print(usbC_VCCVal1_noload)
        print(usbC_VCCVal2_noload)


        #PD value check
        #check 9V,noload，and display the small value.
        if (usbC_9V1_Vcc_noload > PD9V_Vcc_Min_noload) and  (usbC_9V1_Vcc_noload <= PD9V_Vcc_Max_noload):
            if (usbC_9V2_Vcc_noload > PD9V_Vcc_Min_noload) and  (usbC_9V2_Vcc_noload <= PD9V_Vcc_Max_noload):
                usbC_VCCVal_noloadLabel.config(bg = "green")

        else:
            if usbC_9V1_Vcc_noload <= usbC_9V2_Vcc_noload:
                usbC_VCCVal_noloadLabel.config(bg = "red",text='{0:.2f}V'.format(usbC_9V1_Vcc_noload))
                rolling_displayLabel.config(bg="red", text="usbC_VCC1 9V fail_noload")
            else:
                usbC_VCCVal_noloadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_9V2_Vcc_noload))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 9V fail_noload")
            usbTestFailed = 1
            print("9V values no load")
            print(usbC_9V1_Vcc_noload)
            print(usbC_9V2_Vcc_noload)

        #check 15V,noload
        if (usbC_15V1_Vcc_noload > PD15V_Vcc_Min_noload) and (usbC_15V1_Vcc_noload <= PD15V_Vcc_Max_noload):
            if (usbC_15V2_Vcc_noload > PD15V_Vcc_Min_noload) and (usbC_15V2_Vcc_noload <= PD15V_Vcc_Max_noload):
                    usbC_15V_noloadLabel.config(bg="green")

        else:
            if usbC_15V1_Vcc_noload <= usbC_15V2_Vcc_noload:
                usbC_15V_noloadLabel.config(bg = "red",text='{0:.2f}V'.format(usbC_15V1_Vcc_noload))
                rolling_displayLabel.config(bg="red", text="usbC_VCC1 15V fail_noload")
            else:
                usbC_15V_noloadLabel.config(bg="red", text='{0:.2f}V'.format(usbC_15V2_Vcc_noload))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 15V fail_noload")
            usbTestFailed = 1


        #check 20V,noload
        if (usbC_20V1_Vcc_noload > PD20V_Vcc_Min_noload) and (usbC_20V1_Vcc_noload <= PD20V_Vcc_Max_noload):
            if (usbC_20V2_Vcc_noload > PD20V_Vcc_Min_noload) and (usbC_20V2_Vcc_noload <= PD20V_Vcc_Max_noload):
                    usbC_20V_noloadLabel.config(bg="green")

        else:
            if usbC_20V1_Vcc_noload <= usbC_20V2_Vcc_noload:
                usbC_20V_noloadLabel.config(bg = "red",text='{0:.2f}V'.format(usbC_20V1_Vcc_noload))
                rolling_displayLabel.config(bg="red", text="usbC_VCC1 20V fail_noload")
            else:
                usbC_20V_noloadLabel.config(bg="red", text='{0:.2f}V'.format(usbC_20V2_Vcc_noload))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 20V fail_noload")
            usbTestFailed = 1

        # display no load data
        win.update();
        time.sleep(0.1)
        

        ##start with load, turn on both A and C resistor
        test_phase_CC2()
       




        if (usbA_VCCVal > Vcc_Min_load) and  (usbA_VCCVal <Vcc_Max_load):
            usbA_VCCLabel.config(bg = "green")
        else:
            usbA_VCCLabel.config(bg = "red")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text ="usbA_Vcc fail_load");
            
        if (usbA_DMVal >= DM_MIN) and  (usbA_DMVal <= DM_MAX):
            usbA_DMLabel.config(bg = "green")
        else:
            usbA_DMLabel.config(bg = "red")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text ="usbA_DM fail_load");
            
        if (usbA_DPVal >= DP_MIN) and  (usbA_DPVal <= DP_MAX):
            usbA_DPLabel.config(bg = "green")
        else:
            usbA_DPLabel.config(bg = "red")
            usbTestFailed = 1
            rolling_displayLabel.config(bg="red", text ="usbA_DP fail_load");




        # check 5V, load
        if (usbC_VCCVal1_load > Vcc_Min_load) and (usbC_VCCVal1_load <= Vcc_Max_load):
            if (usbC_VCCVal2_load > Vcc_Min_load) and (usbC_VCCVal2_load <= Vcc_Max_load):
                usbC_VCCVal_loadLabel.config(bg="green")
                print("pass")
            else:
                usbC_VCCVal_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_VCCVal2_load))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 fail_load")
                print("Type CCCCCCCCCCCC Vcc2 fail")
                usbTestFailed = 1;

        else:
            usbC_VCCVal_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_VCCVal1_load))
            rolling_displayLabel.config(bg="red", text="usbC_VCC1 fail_load")


            usbTestFailed = 1;
        print("Vcc values")
        print(usbC_VCCVal1_load)
        print(usbC_VCCVal2_load)

        # PD value check
        # check 9V,load
        if (usbC_9V1_Vcc_load > PD9V_Vcc_Min_load) and (usbC_9V1_Vcc_load <= PD9V_Vcc_Max_load):
            if (usbC_9V2_Vcc_load > PD9V_Vcc_Min_load) and (usbC_9V2_Vcc_load <= PD9V_Vcc_Max_load):
                usbC_VCCVal_loadLabel.config(bg="green")
        else:
            print("didn't pass 9V PD load.")
            if usbC_9V1_Vcc_load <= usbC_9V2_Vcc_load:
                usbC_VCCVal_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_9V1_Vcc_load))
                rolling_displayLabel.config(bg="red", text="usbC_VCC1 9V fail_load")

            else:
                usbC_VCCVal_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_9V2_Vcc_load))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 9V fail_load")
            usbTestFailed = 1;
        print("9V values")
        print(usbC_9V1_Vcc_load)
        print(usbC_9V2_Vcc_load)

        # check 15V,load
        if (usbC_15V1_Vcc_load > PD15V_Vcc_Min_load) and (usbC_15V1_Vcc_load <= PD15V_Vcc_Max_load):
            if (usbC_15V2_Vcc_load > PD15V_Vcc_Min_load) and (usbC_15V2_Vcc_load <= PD15V_Vcc_Max_load):
                usbC_15V_loadLabel.config(bg="green")
            #else:
                # print("showing something");
        else:
            if usbC_15V1_Vcc_load <= usbC_15V2_Vcc_load:
                usbC_15V_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_15V1_Vcc_load))
                rolling_displayLabel.config(bg="red", text="usbC_VCC1 15V fail_load")
            else:
                usbC_15V_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_15V2_Vcc_load))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 15V fail_load")
                print("15V with load fail")
            usbTestFailed = 1;
            # print("come on")

        # check 20V,load
        if (usbC_20V1_Vcc_load > PD20V_Vcc_Min_load) and (usbC_20V1_Vcc_load <= PD20V_Vcc_Max_load):
            if (usbC_20V2_Vcc_load > PD20V_Vcc_Min_load) and (usbC_20V2_Vcc_load <= PD20V_Vcc_Max_load):
                usbC_20V_loadLabel.config(bg="green")
        else:
            if usbC_20V1_Vcc_load <= usbC_20V2_Vcc_load:
                usbC_20V_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_20V1_Vcc_load))
                rolling_displayLabel.config(bg="red", text="usbC_VCC1 20V fail_load")
            else:
                usbC_20V_loadLabel.config(bg="red",text='{0:.2f}V'.format(usbC_20V2_Vcc_load))
                rolling_displayLabel.config(bg="red", text="usbC_VCC2 20V fail_load");
            usbTestFailed = 1;

        # display load data
        win.update();
        time.sleep(0.1)


    if (usbTestFailed > 0 ):
        #testButton.config(bg="red", text ="FAIL,click to test")
        testButton.config(bg="red", text ="TEST")
        RetestButton.config(bg=orig_color,text ="RETEST")
        label_passLabel.config(bg="red", text ="FAIL")
        # unitsTestedFAIL += 1
        test_result = 'FAIL'
# detect whether someone scan the barcode
        print("fail")
        
    else:
        #testButton.config(bg="green", text ="PASS,click to test")
        testButton.config(bg=orig_color,text ="TEST")
        RetestButton.config(bg=orig_color,text ="RETEST")
        label_passLabel.config(bg="green", text ="PASS")
        # unitsTestedPASS += 1
        rolling_displayLabel.config(bg="green", text ="pass")
        test_result = 'PASS'

    # infoLabelString = "TESTED " +str(unitsTestedFAIL+unitsTestedPASS) +" PASS " +str(unitsTestedPASS) +" FAIL " +str(unitsTestedFAIL)
    # infoLabel.config(text=infoLabelString)

    # display final result
    win.update();

    # test_number = unitsTestedFAIL + unitsTestedPASS
    #write documents
    # start to write Openpyxl
    global filename
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    reading_time = str(datetime.datetime.now())[10:19]
    A = str(usbA_VCCVal_noload)[:5]
    B = str(usbA_DPVal_noload)[:5]
    C = str(usbA_DMVal_noload)[:5]
    D = str(usbC_VCCzero_noload)[:5]
    E = str(usbC_VCCVal1_noload)[:5]
    F = str(usbC_VCCVal2_noload)[:5]
    G = str(usbC_9V1_Vcc_noload)[:5]
    H = str(usbC_9V2_Vcc_noload)[:5]
    I = str(usbC_15V1_Vcc_noload)[:5]
    J = str(usbC_15V2_Vcc_noload)[:5]
    K = str(usbC_20V1_Vcc_noload)[:5]
    L = str(usbC_20V2_Vcc_noload)[:5]
    M = str(usbA_VCCVal)[:5]
    N = str(usbA_DPVal)[:5]
    O = str(usbA_DMVal)[:5]
    P = str(usbC_VCCzero_load)[:5]
    Q = str(usbC_VCCVal1_load)[:5]
    R = str(usbC_VCCVal2_load)[:5]
    S = str(usbC_9V1_Vcc_load)[:5]
    T = str(usbC_9V2_Vcc_load)[:5]
    U = str(usbC_15V1_Vcc_load)[:5]
    V = str(usbC_15V2_Vcc_load)[:5]
    W = str(usbC_20V1_Vcc_load)[:5]
    X = str(usbC_20V2_Vcc_load)[:5]
    result = test_result


##write down the wb filename into a CSV file

    
    with open('wbnamesAC' + '.csv', 'a') as csv_file:

        writer = csv.writer(csv_file)
        writer.writerow([filename])

# 18. (append the button via config.) 
    Release_buttons()


#release the button when things go wrong, reset buttons.
def Release_buttons():
    testButton.config(state=tk.NORMAL)
    RetestButton.config(state=tk.NORMAL)
    testButton.config(bg=orig_color, text="TEST")
    RetestButton.config(bg=orig_color, text="RETEST")
    label_passLabel.config(bg="green", text="Ready to test")


def WriteExcelforTest():
        #no highlight
        result = test_result
        ws.append(
            [reading_time, MO, Lnum, Pnum, test_number, A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U,
             V, W, X, result])
        wb.save(filename)
        # colorscale()


# for retest, mark "retest" in Excel, and don't change test number.
def WriteExcelforRetest():

        highlight = 'RETEST'
        ws.append(
            [reading_time, MO, Lnum, Pnum, test_number, A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U,
             V, W, X, result,highlight])
        wb.save(filename)
        ##        colorscale()


#need to be fixed and simplify

def exitProgram():
    print("Exit Button pressed")
    win.destroy()



def colorscale():
    from openpyxl.styles import Color, PatternFill, Font, Border
    redFill = PatternFill(start_color="FF0000",end_color="FF0000", fill_type = "solid")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for i in range (2,ws.max_row+1):
##        print(ws.cell(row=i,column=6).value)
        if (float(ws.cell(row=i,column=6).value) > Vcc_Max_noload) or (float(ws.cell(row=i,column=6).value)<Vcc_Min_noload) :
            #print("VCC A NOLOAD")
            ws.cell(row=i,column=6).fill = redFill
        if (float(ws.cell(row=i,column=7).value) > DP_MAX) or (float(ws.cell(row=i,column=7).value)<DP_MIN) :
            #print("A DP NOLOAD")
            ws.cell(row=i,column=7).fill = redFill
        if (float(ws.cell(row=i,column=8).value) > DM_MAX) or (float(ws.cell(row=i,column=8).value)<DM_MIN) :
            #print("A DM NOLOAD")
            ws.cell(row=i,column=8).fill = redFill
        if (float(ws.cell(row=i,column=10).value) > Vcc_Max_noload) or (float(ws.cell(row=i,column=10).value)<Vcc_Min_noload) :
            #print("VCC CC1 NOLOAD")
            ws.cell(row=i,column=10).fill = redFill
        if (float(ws.cell(row=i,column=11).value) > Vcc_Max_noload) or (float(ws.cell(row=i,column=11).value)<Vcc_Min_noload) :
            #print("VCC CC2 NOLOAD")
            ws.cell(row=i,column=11).fill = redFill
        if (float(ws.cell(row=i,column=16).value) > Vcc_Max_load) or (float(ws.cell(row=i,column=16).value)<Vcc_Min_load) :
            #print("VCC A LOAD")
            ws.cell(row=i,column=16).fill = redFill
        if (float(ws.cell(row=i,column=20).value) > Vcc_Max_load) or (float(ws.cell(row=i,column=20).value)<Vcc_Min_load) :
            #print("VCC CC1 LOAD")
            ws.cell(row=i,column=20).fill = redFill
        if (float(ws.cell(row=i,column=21).value) > Vcc_Max_load) or (float(ws.cell(row=i,column=21).value)<Vcc_Min_load) :
            #print("VCC CC2 LOAD")
            ws.cell(row=i,column=21).fill = redFill
        if (float(ws.cell(row=i,column=17).value) > DP_MAX) or (float(ws.cell(row=i,column=17).value)<DM_MIN) :
            #print("A DP LOAD")
            ws.cell(row=i,column=17).fill = redFill
        if (float(ws.cell(row=i,column=18).value) > DM_MAX) or (float(ws.cell(row=i,column=18).value)<DM_MIN) :
            #print("A DM LOAD")
            ws.cell(row=i,column=18).fill = redFill
            
            
        if (float(ws.cell(row=i,column=12).value) > DP_MAX) or (float(ws.cell(row=i,column=12).value)<DP_MIN) :
            #print("C DP1 LOAD")
            ws.cell(row=i,column=12).fill = redFill
        if (float(ws.cell(row=i,column=13).value) > DM_MAX) or (float(ws.cell(row=i,column=13).value)<DM_MIN) :
            #print("C DM1 LOAD")
            ws.cell(row=i,column=13).fill = redFill

        if (float(ws.cell(row=i,column=14).value) > DP_MAX) or (float(ws.cell(row=i,column=14).value)<DP_MIN) :
            #print("C DP2 LOAD")
            ws.cell(row=i,column=14).fill = redFill
        if (float(ws.cell(row=i,column=15).value) > DM_MAX) or (float(ws.cell(row=i,column=15).value)<DM_MIN) :
            #print("C DM2 LOAD")
            ws.cell(row=i,column=15).fill = redFill


        if (float(ws.cell(row=i,column=22).value) > DP_MAX) or (float(ws.cell(row=i,column=22).value)<DP_MIN) :
            #print("C DP1 LOAD")
            ws.cell(row=i,column=22).fill = redFill
        if (float(ws.cell(row=i,column=23).value) > DM_MAX) or (float(ws.cell(row=i,column=23).value)<DM_MIN) :
            #print("C DM1 LOAD")
            ws.cell(row=i,column=23).fill = redFill

        if (float(ws.cell(row=i,column=24).value) > DP_MAX) or (float(ws.cell(row=i,column=24).value)<DP_MIN) :
            #print("C DP2 LOAD")
            ws.cell(row=i,column=24).fill = redFill
        if (float(ws.cell(row=i,column=25).value) > DM_MAX) or (float(ws.cell(row=i,column=25).value)<DM_MIN) :
            #print("C DM2 LOAD")
            ws.cell(row=i,column=25).fill = redFill

        if (float(ws.cell(row=i,column=9).value) > 0.1) or (float(ws.cell(row=i,column=9).value)<-0.1) :
            #print("C 1 ZERO")
            ws.cell(row=i,column=9).fill = redFill
        if (float(ws.cell(row=i,column=19).value) > 0.1) or (float(ws.cell(row=i,column=19).value)<-0.1) :
            #print("C 2 ZERO")
            ws.cell(row=i,column=19).fill = redFill

            
        
            
        
        wb.save(filename)

   
def Scan(event):
    
    global unitsTestedPASS, unitsTestedFAIL, test_number
    global flag
    global time_in_filename
    global MO
    global Lnum
    global Pnum
    global filename
    global displayPDfail_9V 
    from openpyxl.utils import get_column_letter

    

    test_number = 0
    unitsTestedFAIL = 0
    unitsTestedPASS = 0#reset the counter
    flag = 0 # set flag to detect whether barcode is scaned or not

    infoLabelString = "TESTED " +str(unitsTestedFAIL+unitsTestedPASS) +" PASS " +str(unitsTestedPASS) +" FAIL " +str(unitsTestedFAIL)
    infoLabel.config(text=infoLabelString)

    
    entered_text1 = textentry1.get()
    entered_text2 = textentry2.get()
    entered_text3 = textentry3.get()

    entered_text1 = ''.join(entered_text1.split())
    print(str(entered_text1)+' '+str(entered_text2)+' '+str(entered_text3))
    
    MO = entered_text1
    Lnum = entered_text2
    Pnum = entered_text3

# change plan, won't display at the corner.
    together = ('Mnum:' + MO + '\n' + 'Lnum:' + Lnum + '\n' + 'Pnum:' + Pnum )

#    together = ('Mnum:'+MO+'\n'+'Lnum:'+Lnum+'\n'+'Pnum:'+Pnum+'\n'+
#              'PD9V:'+PD9V+'\n'+'PD15V:'+PD15V+'\n'+'PD20V:'+PD20V)

    displaytext = tk.StringVar()
    displaytext.set(together)
   
    label_boxLabel.config(textvariable=displaytext)



    # create a file with headers
    time_in_filename = str(datetime.datetime.now())[:19]
    filename = str(time_in_filename +'_' + MO + '_' + Lnum + '_' + Pnum + '.xlsx')
    wb= openpyxl.Workbook()
    ws = wb.active
    for column in range(1):
            
        ws.append(['Time','MO', 'Lnum', 'Pnum', 'test_number',
                                 'usbA_VCCVal_noload', 'usbA_DPVal_noload', 'usbA_DMVal_noload', 
                                 'usbC_VCCVal_noload_zero reading', 'usbC_VCCVal1_noload', 'usbC_VCCVal2_noload', 
                                 'usbC_9V1_Vcc_noload', 'usbC_9V2_Vcc_noload', 'usbC_15V1_Vcc_noload', 'usbC_15V2_Vcc_noload',
                                'usbC_20V1_Vcc_noload', 'usbC_20V2_Vcc_noload',
                                 'usbA_VCCVal', 'usbA_DPVal', 'usbA_DMVal', 
                                 'usbC_VCCVal_zero reading load', 'usbC_VCCVal1', 'usbC_VCCVal2',
                                 'usbC_9V1_Vcc_load', 'usbC_9V2_Vcc_load', 'usbC_15V1_Vcc_load', 'usbC_15V2_Vcc_load',
                                'usbC_20V1_Vcc_load', 'usbC_20V2_Vcc_load',
                         'Result','highlight'])
        ws.row_dimensions[1].height = 30
      
    #change the width of columns in Excel 
    column = 1
    while column < 19:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = 23
        column += 1
        
    wb.save(filename)
    
    #change the Scan flag to 1
    flag += 1

    #clear all the scan information and set the focus back to first entry label. 
    textentry1.delete(0,'end')
    textentry2.delete(0,'end')
    textentry3.delete(0,'end')
    textentry1.focus()

def ScanOE():
    #this function is to skip scan when sth goes wrong or and there is too much data to store.
    
    global unitsTestedPASS, unitsTestedFAIL, test_number
    global flag
    global time_in_filename
    global MO
    global Lnum
    global Pnum
    global filename
    from openpyxl.utils import get_column_letter

    
    
    test_number = 0
    unitsTestedFAIL = 0
    unitsTestedPASS = 0#reset the counter
    flag = 0 # set flag to detect whether barcode is scaned or not

    infoLabelString = "TESTED " +str(unitsTestedFAIL+unitsTestedPASS) +" PASS " +str(unitsTestedPASS) +" FAIL " +str(unitsTestedFAIL)
    infoLabel.config(text=infoLabelString)

    
    entered_text1 = textentry1.get()
    entered_text2 = textentry2.get()
    entered_text3 = textentry3.get()

    entered_text1 = ''.join(entered_text1.split())
    print(str(entered_text1)+' '+str(entered_text2)+' '+str(entered_text3))
    
    MO = entered_text1
    Lnum = entered_text2
    Pnum = entered_text3
    together = 'Mnum:'+MO+'\n'+'Lnum:'+Lnum+'\n'+'Pnum:'+Pnum
    displaytext = tk.StringVar()
    displaytext.set(together)
   
    label_boxLabel.config(textvariable=displaytext)

# create a file with headers 
    time_in_filename = str(datetime.datetime.now())[:19]
    filename = str(time_in_filename +'_' + MO + '_' + Lnum + '_' + Pnum + '.xlsx')
    wb= openpyxl.Workbook()
    ws = wb.active
    for column in range(1):
        ws.append(['Time','MO', 'Lnum', 'Pnum', 'test_number',
                                 'usbA_VCCVal_noload', 'usbA_DPVal_noload', 'usbA_DMVal_noload', 
                                 'usbC_VCCVal_noload_zero reading', 'usbC_VCCVal1_noload', 'usbC_VCCVal2_noload', 
                                 'usbC_DP1Val_noload', 'usbC_DM1Val_noload', 'usbC_DP2Val_noload', 'usbC_DM2Val_noload', 
                                 'usbA_VCCVal', 'usbA_DPVal', 'usbA_DMVal', 
                                 'usbC_VCCVal_zero reading', 'usbC_VCCVal1', 'usbC_VCCVal2', 
                                 'usbC_DPVal1', 'usbC_DMVal1', 'usbC_DPVal2', 'usbC_DMVal2',
                         'Result','highlight'])
        ws.row_dimensions[1].height = 30
    
        
    column = 1
    while column < 19:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = 23
        column += 1
        
    wb.save(filename)
    flag += 1
    
    textentry1.focus()



def send_email():
    global filename
    
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders
    email_user = 'tuf003tester@gmail.com'
    email_password = 'raspberrypi'
    email_send = 'tuf003tester@gmail.com'#recipient_email

    subject = 'subject'

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send
    msg['Subject'] = subject

    body = 'Hi there, sending this email from Python!'
    msg.attach(MIMEText(body,'plain'))


    attachment  =open(filename,'rb')

    part = MIMEBase('application','octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',"attachment; filename= "+filename)

    msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP('smtp.gmail.com',587)
    server.starttls()
    server.login(email_user,email_password)


    server.sendmail(email_user,email_send,text)
    server.quit()
    
def Restart():
    popupvalue = messagebox.askquestion('WARNING!!!', 'Do you want to restart the program?')
    if popupvalue == 'yes':

        win.destroy()
        # trying to find a way to restart the program under same shell. 
        
        os.system("python3 /home/pi/30.py")
##        exec(open("/home/pi/30.py").read())
        

    
def Restore():
    popupvalue = messagebox.askquestion('WARNING!!!','Are you sure to restore the original settings?')
    if popupvalue == 'yes':
        
        #load the compensation settings 
        with open('defaultsettings' + '.csv', 'r') as csv_file:
            reader = csv.reader(csv_file)
            for column in reader:
                TypeC_cable_chip=(float(column[0]))
                usbC_DM_cable_chip=(float(column[1]))
                usbC_DP_cable_chip=(float(column[2]))
                
                TypeA_cable_chip=(float(column[3]))
                usbA_DM_cable_chip=(float(column[4]))
                usbA_DP_cable_chip=(float(column[5]))



        with open('settings' + '.csv', 'w') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow([TypeC_cable_chip, usbC_DM_cable_chip, usbC_DP_cable_chip, TypeA_cable_chip, usbA_DM_cable_chip, usbA_DP_cable_chip,TypeAA_cable_chip, usbAA_DM_cable_chip, usbAA_DP_cable_chip])

        with open('ACorAA' + '.csv', 'w') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow([0])

            popupvalue = messagebox.askquestion('WARNING!!!','System settings restored, do you want to reboot now?')
            if popupvalue == 'yes':
                os.system("sudo python3 /home/pi/29.py&")
                win.destroy()

        with open('wbnamesAC' + '.csv', 'w') as csv_file:
            pass
        with open('wbnamesAA' + '.csv', 'w') as csv_file:
            pass

       

def Focus(event):
    textentry1.focus()
    print('focus set')



def AutoCalibration():
    popupreturn = messagebox.askquestion('Calibration','Both cables will be calibrated, continue?')
##    print("return:" + popupreturn)
    if popupreturn == 'yes':
        popupvalue = messagebox.askquestion('WARNING!!!','Please make sure cables are connected, press YES to calibrate, system will restart after calibration.')
        
        if popupvalue == 'yes':
            Calibration()               #AC calibration
            os.system("sudo python3 /home/pi/19.py&")
            win.destroy()

##if choose no, exit pop up window.
    if popupreturn == 'no':
        messagebox.askquestion.destroy()
        
    

##AC calibration
def Calibration(n_point=10,t_delta=0.001):

    global TypeAA_cable_chip, usbAA_DM_cable_chip, usbAA_DP_cable_chip

##    
    # readings from the meter when load is connected:
    typeCVccReal = 5.0
    typeCDMReal = 2.5
    typeCDPReal = 2.5
    typeAVccReal = 5
    typeADMReal = 2.44
    typeADPReal = 2.43
    #calculate the calibration
    #measure the ADC readings and take average of the list

    #turn off all ports
    iobus = ExpanderPi.IO()
    iobus.set_port_direction(0, 0x00)
    iobus.write_port(0, 0x00)
    iobus.write_pin(4, 0)
    iobus.write_pin(6, 0)
    iobus.write_pin(7, 0)
    iobus.write_pin(1, 0)
    time.sleep(0.3)
    
##C-port calibration
       

    #turn on two resistors 
    iobus.write_pin(4, 1)
    iobus.write_pin(6, 1) #1.5ohm type C resistor 
    iobus.write_pin(7, 1) #CC2 tpye C
    time.sleep(0.3)
    
    #sample data from type C ports 
    usbC_VCCVal1_sam= [] ;usbC_DMVal1_sam = []; usbC_DPVal1_sam = []
    
    for i in range(n_point):
        usbC_VCCVal1_sam.append(2.0*adc.read_adc_voltage(4, 0))
        usbC_DMVal1_sam.append(2.0*adc.read_adc_voltage(5, 0))
        usbC_DPVal1_sam.append(2.0*adc.read_adc_voltage(6, 0))
        time.sleep(t_delta)

    #omit the highest and the lowest
    usbC_VCCVal1_sam.remove(max(usbC_VCCVal1_sam))
    usbC_VCCVal1_sam.remove(min(usbC_VCCVal1_sam))
    usbC_DMVal1_sam.remove(max(usbC_DMVal1_sam))
    usbC_DMVal1_sam.remove(min(usbC_DMVal1_sam))
    usbC_DPVal1_sam.remove(max(usbC_DPVal1_sam))
    usbC_DPVal1_sam.remove(min(usbC_DPVal1_sam))

    # take average
    usbC_VCCVal1 = sum(usbC_VCCVal1_sam)/ len(usbC_VCCVal1_sam)
    usbC_DMVal1 = sum(usbC_DMVal1_sam)/ len(usbC_DMVal1_sam)
    usbC_DPVal1 = sum(usbC_DPVal1_sam)/ len(usbC_DPVal1_sam)
    

    #calculate the offset 
    TypeC_cable_chip = typeCVccReal - (usbC_VCCVal1)
    usbC_DM_cable_chip = typeCDMReal - (usbC_DMVal1)
    usbC_DP_cable_chip = typeCDPReal - (usbC_DPVal1)

    
####A-port calibration
##    iobus.write_pin(4, 0)
##    iobus.write_pin(6, 0)
##    iobus.write_pin(7, 0)
##    iobus.write_pin(1, 0)
##    time.sleep(0.1)
##    #turn on two resistors 
##    iobus.write_pin(4, 1)
##    iobus.write_pin(6, 1)
##    time.sleep(0.2)
    #sample data from type A ports
    usbA_VCCVal_sam = []; usbA_DMVal_sam = []; usbA_DPVal_sam = []
    
    for i in range(n_point):
        usbA_VCCVal_sam.append(2.0*(adc.read_adc_voltage(1, 0)))
        usbA_DMVal_sam.append(2.0*adc.read_adc_voltage(2, 0))
        usbA_DPVal_sam.append(2.0*adc.read_adc_voltage(3, 0))
        time.sleep(t_delta)
    print(usbA_VCCVal_sam)
        
    
    #omit the highest and the lowest
    usbA_VCCVal_sam.remove(max(usbA_VCCVal_sam))
    usbA_VCCVal_sam.remove(min(usbA_VCCVal_sam))
    usbA_DMVal_sam.remove(max(usbA_DMVal_sam))
    usbA_DMVal_sam.remove(min(usbA_DMVal_sam))
    usbA_DPVal_sam.remove(max(usbA_DPVal_sam))
    usbA_DPVal_sam.remove(min(usbA_DPVal_sam))
    print(usbA_VCCVal_sam)
    # take average
    usbA_VCCVal = sum(usbA_VCCVal_sam)/ len(usbA_VCCVal_sam)
    usbA_DMVal = sum(usbA_DMVal_sam)/ len(usbA_DMVal_sam)
    usbA_DPVal = sum(usbA_DPVal_sam)/ len(usbA_DPVal_sam)

    print(usbA_VCCVal)   
    #calculate the offset
    TypeA_cable_chip = typeAVccReal - (usbA_VCCVal)
    usbA_DM_cable_chip = typeADMReal - (usbA_DMVal)
    usbA_DP_cable_chip = typeADPReal - (usbA_DPVal)

    print(TypeA_cable_chip)
    
    #write the calibration settings into a CSV file.
    
    with open('settings' + '.csv', 'w') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow([TypeC_cable_chip, usbC_DM_cable_chip, usbC_DP_cable_chip, TypeA_cable_chip, usbA_DM_cable_chip, usbA_DP_cable_chip,
                             TypeAA_cable_chip, usbAA_DM_cable_chip, usbAA_DP_cable_chip])

    print('typeC: '+ str(TypeC_cable_chip), str(usbC_DM_cable_chip),
                  str(usbC_DP_cable_chip), '\n',
                  'typeA: '+str(TypeA_cable_chip), str(usbA_DM_cable_chip),
                  str(usbA_DP_cable_chip), '\n',
              'typeA: '+str(TypeAA_cable_chip), str(usbAA_DM_cable_chip),
                  str(usbAA_DP_cable_chip))

    iobus.write_pin(4, 0)
    iobus.write_pin(6, 0)
    iobus.write_pin(7, 0)
    iobus.write_pin(1, 0)

    
def TUF005():
    global testButton
    with open('TUF4or5' + '.csv', 'w') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow([0])
    print('tuf0055555')
    messagebox.showinfo(title='TUF005',message='            TUF005 set              ')
    Restart()

def TUF004():
    global testButton
    with open('TUF4or5' + '.csv', 'w') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow([1])
    print('yes tuf004444')
    messagebox.showinfo(title='TUF004', message='            TUF004 set              ')
    Restart()


def newfunc():
    global window
    #create a pop up window and put it in the middle,first 2 digits are size, last 2 represent the position
    window = tk.Toplevel(win)
    window.geometry("%dx%d%+d%+d" % (350, 200, 230, 100))
    window.title("function box")


    Twidth = 12
    #send email in new function
    #Send_emailButton  = tk.Button(window, text = "Send Email", font = "Helvetica 16 bold",
    #                    command = send_email, height = 2 , width = Twidth, borderwidth=2)
    #Send_emailButton.grid(row=1,column=0)

    #calibrate cable compensation for cable replacement
    CalibrationButton  = tk.Button(window, text = "Auto-Calibrate", font = "Helvetica 16 bold",
                        command = AutoCalibration, height = 2 , width = Twidth, borderwidth=2)
    CalibrationButton.grid(row=1,column=1)

    # restart
    RestartButton = tk.Button(window, text="Restart", font="Helvetica 16 bold",
                              command=Restart, height=2, width=Twidth, borderwidth=2)
    RestartButton.grid(row=1, column=0)

    #restore calibration settings
    RestoreButton  = tk.Button(window, text = "Restore settings", font = "Helvetica 16 bold",
                        command = Restore, height = 2 , width = Twidth, borderwidth=2)
    RestoreButton.grid(row=2,column=1)

    # release test and retest buttons
    ReleaseButton = tk.Button(window, text="Release Button", font="Helvetica 16 bold",
                              command=Release_buttons, height=2, width=Twidth, borderwidth=2)
    ReleaseButton.grid(row=2, column=0)

    # change to test TUF004
    TUF004Button = tk.Button(window, text="Test TUF004", font="Helvetica 16 bold",
                             command=TUF004, height=2, width=Twidth, borderwidth=2)
    TUF004Button.grid(row=3, column=0)

    # change to test TUF005
    TUF005Button = tk.Button(window, text="Test TUF005", font="Helvetica 16 bold",
                              command=TUF005, height=2, width=Twidth, borderwidth=2)
    TUF005Button.grid(row=3, column=1)


    



     
win.title("ELSAFE/OE QC TUF TESTER")
win.geometry('800x480')

# set the reference voltage.  this should be set to the exact voltage
# measured on the Expander Pi Vref pin.
adc.set_adc_refvoltage(4.096)
    
##exitButton  = tk.Button(win, text = "Exit", font = "Helvetica 16 bold",
##                        command = exitProgram, height = 1 , width = 11,borderwidth=2)
##exitButton.grid(row=2,column=4)

FONT = "Helvetica 13 bold"
WIDTH = 19
WIDTH1 = 16
HEIGHT = 3

# decide whether to test TUF004 or TUF005, 1 for TUF004, 0 for TUF005

with open('TUF4or5' + '.csv', 'r') as csv_file:
    reader = csv.reader(csv_file)
    for column in reader:
        TUFflag = (int(column[0]))
        if TUFflag == 0 or TUFflag == 1:
            if TUFflag == 1:
                print('TUF004 mode')
            if TUFflag == 0:
                print('TUF005 mode')
        else:
            print('CSV ERROR');
        print("TUFflag:")
        print(TUFflag)


usb_VCCLabel = tk.Label(win, text='VCC', height = 2, width = 12, borderwidth=1, bg = "light blue", font = "Helvetica 16 bold")
usb_VCCLabel.grid(row=0, column = 1)

usb_DMLabel = tk.Label(win, text='DP', height = 2, width = 12, borderwidth=1, bg ="light blue", font = "Helvetica 16 bold")
usb_DMLabel.grid(row=0, column = 3)

usb_DPLabel = tk.Label(win, text='DM', height = 2, width = 12, borderwidth=1, bg = "light blue", font = "Helvetica 16 bold")
usb_DPLabel.grid(row=0, column = 2)


usb1_Label = tk.Label(win, text='USB A 2A', height = HEIGHT, width = WIDTH, borderwidth=1, bg = "dark orange", font = FONT)
usb1_Label.grid(row=2, column = 0)
usbA_VCCLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbA_VCCLabel.grid(row=2, column = 1)
usbA_DMLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg ="green", font = FONT)
usbA_DMLabel.grid(row=2, column = 2)
usbA_DPLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbA_DPLabel.grid(row=2, column = 3)

usb3_Label = tk.Label(win, text='USB A 0A', height = HEIGHT, width = WIDTH, borderwidth=1, bg = "dark orange", font = FONT)
usb3_Label.grid(row=3, column = 0)
usbA_VCCVal_noloadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbA_VCCVal_noloadLabel.grid(row=3, column = 1)
usbA_DMVal_noloadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg ="green", font = FONT)
usbA_DMVal_noloadLabel.grid(row=3, column = 2)
usbA_DPVal_noloadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbA_DPVal_noloadLabel.grid(row=3, column = 3)


##type C doesn't need DP DM as there are changing

usb2_Label = tk.Label(win, text='USB C 3A', height = HEIGHT, width = WIDTH, borderwidth=1, bg = "light blue", font = FONT)
usb2_Label.grid(row=4, column = 0)


###!!! in the new setup, check zero and 5V and 9V will be all shown on the same spot, with same label.
usbC_VCCVal_loadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbC_VCCVal_loadLabel.grid(row=4, column = 1)
usbC_15V_loadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbC_15V_loadLabel.grid(row=4, column = 2)
usbC_20V_loadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbC_20V_loadLabel.grid(row=4, column = 3, sticky=tk.W)

usb4_Label = tk.Label(win, text='USB C 0A', height = HEIGHT, width = WIDTH, borderwidth=1, bg = "light blue", font = FONT)
usb4_Label.grid(row=5, column = 0)
#check zero output #check5V,9V
usbC_VCCVal_noloadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbC_VCCVal_noloadLabel.grid(row=5, column = 1)
usbC_15V_noloadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbC_15V_noloadLabel.grid(row=5, column = 2)
usbC_20V_noloadLabel = tk.Label(win, text="", height = HEIGHT, width = WIDTH1, borderwidth=1, bg = "green", font = FONT)
usbC_20V_noloadLabel.grid(row=5, column = 3, sticky=tk.W)



rolling_displayLabel = tk.Label(win, text="",width = 25, borderwidth=1, bg = "green", font = "Helvetica 13 bold")
rolling_displayLabel.grid(row=7, column = 0, columnspan=3,sticky=tk.W)

##create a frame to contain 3 input text entries
##right_frame = Frame(win, bg='gray', width=150, height=150)
##right_frame.grid(row=3,column=4,rowspan=3,sticky=tk.N)

#labelframe
labelframe1 = tk.LabelFrame(win,height=50,width=170,labelanchor='nw',text='Mnum',bd=0)
labelframe1.grid(row=3,column=4)
labelframe1.visible=False
labelframe2 = tk.LabelFrame(win,height=50,width=170,labelanchor='nw',text='Lnum',bd=0)
labelframe2.grid(row=4,column=4)
labelframe3 = tk.LabelFrame(win,height=50,width=170,labelanchor='nw',text='Pnum',bd=0)
labelframe3.grid(row=5,column=4)

##labelframe4.visible=False
# create a text box to display 3 lines
#output = tk.Text(win, height = 6, width = 26, borderwidth=1, bg = "white", font = "Helvetica 11 bold")
#output.grid(row=6, column = 3, columnspan=3)

# create a label box to display 3 lines
label_boxLabel = tk.Label(win, text="", height = 8, width = 20, bg='white', borderwidth=1, font = "Helvetica 12 bold")
label_boxLabel.grid(row=6, column = 3, columnspan=2,sticky = tk.E)

# PASS or not label box
label_passLabel = tk.Label(win, text="", height = 8, width = 15, bg='white', borderwidth=1, font = "Helvetica 12 bold")
label_passLabel.grid(row=6, column = 3, columnspan=1,sticky = tk.W)

#create a text entry1 
textentry1 = tk.Entry(win, width=20, borderwidth=1, bg='white')
textentry1.grid(row=3,column=4)
textentry1.focus()


#create a text entry2 
textentry2 = tk.Entry(win, width=20, borderwidth=1, bg='white')
textentry2.grid(row=4,column=4)



#create a text entry3 
textentry3 = tk.Entry(win, width=20, borderwidth=1, bg='white')
textentry3.grid(row=5,column=4)
# bind the text entry, so that it will lead to Scan(). this is done by have <space> or <Tab> in the bar code.
textentry3.bind('<Tab>',Scan)
textentry3.bind('<space>',Scan)

##Send_emailButton  = tk.Button(win, text = "Finish", font = "Helvetica 16 bold",
##                        command = send_email, height = 1 , width = 11,borderwidth=2)
##Send_emailButton.grid(row=0,column=4)


# add a new window
FunctionButton  = tk.Button(win, text = "Functions", font = "Helvetica 16 bold",
                        command = newfunc, height = 1 , width = 12, borderwidth=2)
FunctionButton.grid(row=0,column=4, sticky=tk.W)


#next job button, skip scan
NextJobButton  = tk.Button(win, text = "Skip scan", font = "Helvetica 16 bold",
                        command = ScanOE, height = 1 , width = 12, borderwidth=2)
NextJobButton.grid(row=2,column=4, sticky=tk.W)



TUFLabel = tk.Label(win, text="TUF", font = "Helvetica 13",)
TUFLabel.grid(row=7, column=2)


infoLabel = tk.Label(win, text="TESTED 0, PASS 0, FAIL 0", font = "Helvetica 13",)
infoLabel.grid(row=7, column=3, columnspan=3)

gif = tk.PhotoImage(file="/home/pi/elsafeicon.gif")
gif = gif.subsample(17,17)


iconLabel = tk.Label(win, image=gif)
iconLabel.grid(row=0, column=0, columnspan=1, sticky = tk.W)


# test either TUF004 or TUF005.
if TUFflag == 1:
    testButton = tk.Button(win, text="TEST TUF004", command=qcTestUSB,
                           font="Helvetica 18 bold",
                           height=5, width=19, borderwidth=2)

    testButton.grid(row=6, column=0, columnspan=2, sticky=tk.W)
    orig_color = testButton.cget("background")

    testButton.bind('<FocusIn>', Focus)

    RetestButton = tk.Button(win, text="RETEST004", font="Helvetica 18 bold",
                             command=Retest, height=5, width=13)

    RetestButton.grid(row=6, column=1, columnspan=2, sticky=tk.E)
    TUFLabel.config(text="TUF004")
else:
    testButton = tk.Button(win, text="TEST TUF005", command=qcTestUSB,
                           font="Helvetica 18 bold",
                           height=5, width=19, borderwidth=2)

    testButton.grid(row=6, column=0, columnspan=2, sticky=tk.W)
    orig_color = testButton.cget("background")

    testButton.bind('<FocusIn>', Focus)

    RetestButton = tk.Button(win, text="RETEST005", font="Helvetica 18 bold",
                             command=Retest, height=5, width=13)

    RetestButton.grid(row=6, column=1, columnspan=2, sticky=tk.E)
    TUFLabel.config(text="TUF005")

win.mainloop()

    
 

         
       

    
              
        

    
        
